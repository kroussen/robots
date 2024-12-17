import json
import random
from datetime import datetime, timedelta

from django.db.models import Count
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.utils.dateparse import parse_datetime
from django.utils.decorators import method_decorator
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Robot


# Create your views here.
@method_decorator(csrf_exempt, name='dispatch')
class RobotCreateView(View):

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body.decode('utf-8'))

            model = data.get('model')
            version = data.get('version')
            created = data.get('created')

            if not model or not version or not created:
                return JsonResponse(
                    {"message": "Ошибка: Поля 'model', 'version' и 'created' обязательны"},
                    status=400
                )

            parsed_created = parse_datetime(created)
            if not parsed_created:
                return JsonResponse(
                    {"message": "Ошибка: Некорректный формат даты (ожидается 'YYYY-MM-DD HH:MM:SS')"},
                    status=400
                )

            serial = self.generate_serial()
            robot = Robot(serial=serial, model=model, version=version, created=parsed_created)
            robot.full_clean()
            robot.save()

            return JsonResponse(
                {"message": "Робот успешно создан",
                 "data": {"serial": serial, "model": model, "version": version, "created": created}},
                status=201
            )

        except json.JSONDecodeError:
            return JsonResponse({"message": "Ошибка: Некорректный формат JSON"}, status=400)

        except ValidationError as e:
            return JsonResponse({"message": "Ошибка валидации данных", "errors": e.message_dict}, status=400)

        except Exception as e:
            return JsonResponse({"message": "Внутренняя ошибка сервера", "error": str(e)}, status=500)

    def generate_serial(self):
        return ''.join([str(random.randint(0, 9)) for _ in range(5)])


def generate_weekly_report(request):
    today = datetime.today()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    robots = Robot.objects.filter(created__range=[start_of_week, end_of_week]) \
        .values('model', 'version') \
        .annotate(total=Count('id')) \
        .order_by('model', 'version')

    wb = Workbook()

    models = set([robot['model'] for robot in robots])

    if models:
        for model in models:
            sheet = wb.create_sheet(title=model)

            sheet.append(["Модель", "Версия", "Количество за неделю"])

            model_data = [robot for robot in robots if robot['model'] == model]

            for data in model_data:
                sheet.append([data['model'], data['version'], data['total']])

            for col in range(1, sheet.max_column + 1):
                max_length = 0
                column = get_column_letter(col)
                for row in sheet.iter_rows():
                    try:
                        cell_value = str(row[col - 1].value)
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    else:

        wb.create_sheet(title="No Data Available")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename="robot_production_report_{start_of_week.strftime("%Y-%m-%d")}_to_{end_of_week.strftime("%Y-%m-%d")}.xlsx"'

    wb.save(response)

    return response
